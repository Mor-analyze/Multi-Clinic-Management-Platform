from fastapi import APIRouter, Depends, HTTPException, Query , UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from pydantic import BaseModel
from datetime import date, datetime
import uuid
import openpyxl
import io

from app.models.database import get_db
from app.models.expense import Expense, ExpenseSubcategory, ExpenseCategory
from app.models.user import User
from app.services.dependencies import get_current_user, require_admin

router = APIRouter(prefix="/expenses", tags=["Expenses"])

class SubcategoryCreate(BaseModel):
    clinic_id: str
    category: str
    name: str

class ExpenseCreate(BaseModel):
    clinic_id: str
    branch_id: str
    subcategory_id: str
    amount: float
    description: Optional[str] = None
    expense_date: date
    reference_num: Optional[str] = None

class OverrideData(BaseModel):
    override_reason: str

@router.get("/subcategories")
def list_subcategories(
    clinic_id: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    subs = db.query(ExpenseSubcategory).filter(
        ExpenseSubcategory.clinic_id == uuid.UUID(clinic_id),
        ExpenseSubcategory.is_active == True
    ).order_by(ExpenseSubcategory.category, ExpenseSubcategory.name).all()

    return [{
        "id": str(s.id),
        "category": s.category.value,
        "name": s.name,
    } for s in subs]

@router.post("/subcategories")
def create_subcategory(
    data: SubcategoryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    sub = ExpenseSubcategory(
        clinic_id=uuid.UUID(data.clinic_id),
        category=ExpenseCategory(data.category),
        name=data.name,
        is_active=True,
    )
    db.add(sub)
    db.commit()
    return {"message": "Subcategory created", "id": str(sub.id)}

@router.post("/")
def create_expense(
    data: ExpenseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    # Check duplicate reference number
    if data.reference_num:
        existing = db.query(Expense).filter(
            Expense.clinic_id == uuid.UUID(data.clinic_id),
            Expense.reference_num == data.reference_num,
            Expense.duplicate_override == False
        ).first()

        if existing:
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "Duplicate reference number detected",
                    "existing_expense": {
                        "date": str(existing.expense_date),
                        "amount": float(existing.amount),
                        "description": existing.description,
                    },
                    "requires_override": True
                }
            )

    expense = Expense(
        clinic_id=uuid.UUID(data.clinic_id),
        branch_id=uuid.UUID(data.branch_id),
        subcategory_id=uuid.UUID(data.subcategory_id),
        amount=data.amount,
        description=data.description,
        expense_date=data.expense_date,
        reference_num=data.reference_num,
        recorded_by=current_user.id,
    )
    db.add(expense)
    db.commit()

    return {"message": "Expense recorded", "id": str(expense.id)}

@router.get("/")
def list_expenses(
    clinic_id: str = Query(...),
    category: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    query = db.query(Expense, ExpenseSubcategory).join(
        ExpenseSubcategory,
        ExpenseSubcategory.id == Expense.subcategory_id
    ).filter(Expense.clinic_id == uuid.UUID(clinic_id))

    if category:
        query = query.filter(
            ExpenseSubcategory.category == ExpenseCategory(category)
        )
    if date_from:
        query = query.filter(Expense.expense_date >= date_from)
    if date_to:
        query = query.filter(Expense.expense_date <= date_to)

    results = query.order_by(
        Expense.expense_date.desc()
    ).offset(skip).limit(limit).all()

    return [{
        "id": str(e.id),
        "category": sub.category.value,
        "subcategory": sub.name,
        "amount": float(e.amount),
        "description": e.description,
        "expense_date": str(e.expense_date),
        "reference_num": e.reference_num,
    } for e, sub in results]

@router.get("/summary")
def expense_summary(
    clinic_id: str = Query(...),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    query = db.query(
        ExpenseSubcategory.category,
        func.sum(Expense.amount)
    ).join(
        ExpenseSubcategory,
        ExpenseSubcategory.id == Expense.subcategory_id
    ).filter(Expense.clinic_id == uuid.UUID(clinic_id))

    if date_from:
        query = query.filter(Expense.expense_date >= date_from)
    if date_to:
        query = query.filter(Expense.expense_date <= date_to)

    results = query.group_by(ExpenseSubcategory.category).all()

    summary = {"fixed": 0, "operations": 0, "office": 0, "total": 0}
    for category, total in results:
        summary[category.value] = float(total or 0)
        summary["total"] += float(total or 0)

    return summary


@router.post("/import")
async def import_expenses_excel(
    file: UploadFile = File(...),
    clinic_id: str = Query(...),
    branch_id: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

    # Find the right sheet
    sheet_name = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "_" not in name and "guide" not in nl and "list" not in nl:
            sheet_name = name
            break

    if not sheet_name and wb.sheetnames:
        sheet_name = wb.sheetnames[0]

    if not sheet_name:
        raise HTTPException(status_code=400, detail="Could not find Expenses sheet in the file")

    ws = wb[sheet_name]
    # Find header row
    headers = {}
    header_row = None
    for row in ws.iter_rows():
        row_values = [str(c.value).strip().lower() if c.value else "" for c in row]
        if any("date" in v for v in row_values) and any("amount" in v for v in row_values):
            header_row = row[0].row
            # Map headers
            for cell in row:
                if cell.value:
                    h = str(cell.value).strip().lower().replace("*", "").strip()
                    if h == "date": headers["date"] = cell.column
                    elif h == "category": headers["category"] = cell.column
                    elif h == "subcategory": headers["subcategory"] = cell.column
                    elif "amount" in h: headers["amount"] = cell.column
                    elif h == "description": headers["description"] = cell.column
                    elif "reference" in h: headers["reference"] = cell.column
            break

    if not header_row:
        raise HTTPException(status_code=400, detail="Could not find header row in the file")

    created = 0
    skipped = 0
    errors = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        try:
            # Skip empty rows
            if not any(row):
                continue

            # Get values
            date_val = row[headers["date"] - 1] if "date" in headers else None
            category_val = str(row[headers["category"] - 1]).strip().lower() if "category" in headers and row[headers["category"] - 1] else None
            subcategory_val = str(row[headers["subcategory"] - 1]).strip() if "subcategory" in headers and row[headers["subcategory"] - 1] else None
            amount_val = row[headers["amount"] - 1] if "amount" in headers else None
            description_val = str(row[headers["description"] - 1]).strip() if "description" in headers and row[headers["description"] - 1] else None
            reference_val = str(row[headers["reference"] - 1]).strip() if "reference" in headers and row[headers["reference"] - 1] else None

            # Skip if missing required fields
            if not date_val or not category_val or not amount_val:
                skipped += 1
                continue

            # Skip total row
            if str(date_val).strip().upper() in ["TOTAL", "TOTAL EXPENSES"]:
                continue

            # Parse date
            if hasattr(date_val, 'date'):
                expense_date = date_val.date()
            else:
                for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"]:
                    try:
                        expense_date = datetime.strptime(str(date_val).strip(), fmt).date()
                        break
                    except:
                        continue
                else:
                    skipped += 1
                    continue

            # Parse amount
            try:
                amount = float(str(amount_val).replace("$", "").replace(",", "").strip())
                if amount <= 0:
                    skipped += 1
                    continue
            except:
                skipped += 1
                continue

            # Map category
            cat_map = {"fixed": "fixed", "operations": "operations", "office": "office", "office & admin": "office", "office and admin": "office"}
            category = cat_map.get(category_val.lower(), None)
            if not category:
                errors.append(f"Unknown category: {category_val}")
                skipped += 1
                continue

            # Find or create subcategory
            sub = None
            if subcategory_val:
                sub = db.query(ExpenseSubcategory).filter(
                    ExpenseSubcategory.clinic_id == uuid.UUID(clinic_id),
                    ExpenseSubcategory.name.ilike(subcategory_val),
                ).first()

                if not sub:
                    sub = ExpenseSubcategory(
                        clinic_id=uuid.UUID(clinic_id),
                        category=ExpenseCategory(category),
                        name=subcategory_val,
                        is_active=True,
                    )
                    db.add(sub)
                    db.flush()

            if not sub:
                skipped += 1
                errors.append(f"Missing subcategory for row with amount {amount}")
                continue

            # Check duplicate reference
            if reference_val:
                existing = db.query(Expense).filter(
                    Expense.clinic_id == uuid.UUID(clinic_id),
                    Expense.reference_num == reference_val,
                    Expense.duplicate_override == False
                ).first()
                if existing:
                    skipped += 1
                    errors.append(f"Duplicate reference: {reference_val}")
                    continue

            expense = Expense(
                clinic_id=uuid.UUID(clinic_id),
                branch_id=uuid.UUID(branch_id),
                subcategory_id=sub.id,
                amount=amount,
                description=description_val,
                expense_date=expense_date,
                reference_num=reference_val,
                recorded_by=current_user.id,
            )
            db.add(expense)
            created += 1

        except Exception as e:
            errors.append(str(e))
            skipped += 1

    db.commit()
    return {
        "message": "Import complete",
        "created": created,
        "skipped": skipped,
        "errors": errors[:10]
    }