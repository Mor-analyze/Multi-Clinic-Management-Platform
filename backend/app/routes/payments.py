from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
import openpyxl
import io
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from pydantic import BaseModel
from datetime import date, datetime
import uuid

from app.models.database import get_db
from app.models.payment import PatientPayment, PaymentAllocation, PaymentMethodType
from app.models.invoice import Invoice, InvoiceStatus
from app.models.patient import Patient
from app.models.user import User
from app.services.dependencies import get_current_user, require_admin

router = APIRouter(prefix="/payments", tags=["Payments"])

class PaymentCreate(BaseModel):
    clinic_id: str
    patient_id: str
    amount: float
    method: str
    reference_num: Optional[str] = None
    payment_date: datetime
    notes: Optional[str] = None

class DuplicateOverride(BaseModel):
    override_reason: str

def apply_fifo(payment: PatientPayment, db: Session):
    remaining = float(payment.amount)
    unpaid_invoices = db.query(Invoice).filter(
        Invoice.patient_id == payment.patient_id,
        Invoice.balance > 0
    ).order_by(Invoice.purchase_date.asc()).all()

    for invoice in unpaid_invoices:
        if remaining <= 0:
            break
        invoice_balance = float(invoice.balance)
        apply_amount = min(remaining, invoice_balance)
        invoice.collected = float(invoice.collected) + apply_amount
        invoice.balance = float(invoice.balance) - apply_amount
        if invoice.balance <= 0:
            invoice.balance = 0
            invoice.status = InvoiceStatus.paid
        else:
            invoice.status = InvoiceStatus.partial

        allocation = PaymentAllocation(
            payment_id=payment.id,
            invoice_id=invoice.id,
            amount_applied=apply_amount,
        )
        db.add(allocation)
        remaining -= apply_amount

    db.commit()

@router.post("/")
def record_payment(
    data: PaymentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Check for duplicate reference number
    if data.reference_num:
        existing = db.query(PatientPayment).filter(
            PatientPayment.clinic_id == uuid.UUID(data.clinic_id),
            PatientPayment.reference_num == data.reference_num,
            PatientPayment.duplicate_override == False
        ).first()

        if existing:
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "Duplicate reference number detected",
                    "existing_payment": {
                        "date": str(existing.payment_date),
                        "amount": float(existing.amount),
                        "method": existing.method.value,
                    },
                    "requires_override": True
                }
            )

    payment = PatientPayment(
        clinic_id=uuid.UUID(data.clinic_id),
        patient_id=uuid.UUID(data.patient_id),
        amount=data.amount,
        method=PaymentMethodType(data.method),
        reference_num=data.reference_num,
        payment_date=data.payment_date,
        notes=data.notes,
        recorded_by=current_user.id,
    )
    db.add(payment)
    db.flush()

    apply_fifo(payment, db)

    return {
        "message": "Payment recorded and applied",
        "payment_id": str(payment.id),
        "amount": data.amount,
        "method": data.method,
    }

@router.post("/{payment_id}/override")
def override_duplicate(
    payment_id: str,
    data: DuplicateOverride,
    original_payment: PaymentCreate = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    payment = PatientPayment(
        clinic_id=uuid.UUID(original_payment.clinic_id),
        patient_id=uuid.UUID(original_payment.patient_id),
        amount=original_payment.amount,
        method=PaymentMethodType(original_payment.method),
        reference_num=original_payment.reference_num,
        payment_date=original_payment.payment_date,
        notes=original_payment.notes,
        duplicate_override=True,
        override_reason=data.override_reason,
        override_by=current_user.id,
        override_at=datetime.utcnow(),
        recorded_by=current_user.id,
    )
    db.add(payment)
    db.flush()
    apply_fifo(payment, db)

    return {"message": "Payment recorded with override", "payment_id": str(payment.id)}

@router.get("/patient/{patient_id}")
def get_patient_payments(
    patient_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    payments = db.query(PatientPayment).filter(
        PatientPayment.patient_id == uuid.UUID(patient_id)
    ).order_by(PatientPayment.payment_date.desc()).all()

    return [{
        "id": str(p.id),
        "amount": float(p.amount),
        "method": p.method.value,
        "reference_num": p.reference_num,
        "payment_date": str(p.payment_date),
        "notes": p.notes,
        "duplicate_override": p.duplicate_override,
    } for p in payments]
    
    
    
@router.post("/import")
async def import_payments_excel(
    file: UploadFile = File(...),
    clinic_id: str = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() in ["payments", "sheet1"] and "sample" not in name.lower() and "guide" not in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        raise HTTPException(status_code=400, detail="Could not find Payments sheet")

    ws = wb[sheet_name]

    # Find headers
    headers = {}
    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip().lower() in ["date", "patient jane id", "amount", "method", "reference #", "notes"]:
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        raise HTTPException(status_code=400, detail="Could not find header row")

    for cell in ws[header_row]:
        if cell.value:
            h = str(cell.value).strip().lower()
            if h == "date": headers["date"] = cell.column
            elif h in ["patient jane id", "patient id"]: headers["patient_id"] = cell.column
            elif h in ["amount", "amount (cad)"]: headers["amount"] = cell.column
            elif h == "method": headers["method"] = cell.column
            elif h in ["reference #", "reference"]: headers["reference"] = cell.column
            elif h == "notes": headers["notes"] = cell.column

    created = 0
    skipped = 0
    errors = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        try:
            if not any(row):
                continue

            date_val = row[headers["date"] - 1] if "date" in headers else None
            jane_id = str(row[headers["patient_id"] - 1]).strip() if "patient_id" in headers and row[headers["patient_id"] - 1] else None
            amount_val = row[headers["amount"] - 1] if "amount" in headers else None
            method_val = str(row[headers["method"] - 1]).strip().lower() if "method" in headers and row[headers["method"] - 1] else None
            reference_val = str(row[headers["reference"] - 1]).strip() if "reference" in headers and row[headers["reference"] - 1] else None
            notes_val = str(row[headers["notes"] - 1]).strip() if "notes" in headers and row[headers["notes"] - 1] else None

            if not date_val or not jane_id or not amount_val or not method_val:
                skipped += 1
                continue

            # Skip total row
            if str(date_val).strip().upper() in ["TOTAL", "TOTAL PAYMENTS"]:
                continue

            # Parse date
            from datetime import datetime
            if hasattr(date_val, 'date'):
                payment_date = datetime.combine(date_val.date(), datetime.min.time())
            else:
                for fmt in ["%Y-%m-%d", "%m/%d/%Y"]:
                    try:
                        payment_date = datetime.strptime(str(date_val).strip(), fmt)
                        break
                    except:
                        continue
                else:
                    skipped += 1
                    continue

            # Parse amount
            try:
                amount = float(str(amount_val).replace("$", "").replace(",", "").strip())
                if amount <= 0:
                    skipped += 1
                    continue
            except:
                skipped += 1
                continue

            # Find patient by Jane ID
            from app.models.patient import Patient
            patient = db.query(Patient).filter(
                Patient.jane_id == jane_id,
                Patient.clinic_id == uuid.UUID(clinic_id)
            ).first()

            if not patient:
                errors.append(f"Patient not found: {jane_id}")
                skipped += 1
                continue

            # Check duplicate reference
            if reference_val:
                existing = db.query(PatientPayment).filter(
                    PatientPayment.clinic_id == uuid.UUID(clinic_id),
                    PatientPayment.reference_num == reference_val,
                    PatientPayment.duplicate_override == False
                ).first()
                if existing:
                    errors.append(f"Duplicate reference: {reference_val}")
                    skipped += 1
                    continue

            # Validate method
            method_map = {"cash": "cash", "etransfer": "etransfer", "e-transfer": "etransfer", "e transfer": "etransfer"}
            method = method_map.get(method_val.lower())
            if not method:
                errors.append(f"Unknown method: {method_val}")
                skipped += 1
                continue

            payment = PatientPayment(
                clinic_id=uuid.UUID(clinic_id),
                patient_id=patient.id,
                amount=amount,
                method=PaymentMethodType(method),
                reference_num=reference_val,
                payment_date=payment_date,
                notes=notes_val,
                recorded_by=current_user.id,
            )
            db.add(payment)
            db.flush()
            apply_fifo(payment, db)
            created += 1

        except Exception as e:
            errors.append(str(e))
            skipped += 1

    db.commit()
    return {
        "message": "Import complete",
        "created": created,
        "skipped": skipped,
        "errors": errors[:10]
    }